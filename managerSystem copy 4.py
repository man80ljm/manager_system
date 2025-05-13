import sys
import sqlite3
from datetime import datetime
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QScrollArea,
                             QLineEdit, QDialog, QMessageBox, QFormLayout,
                             QInputDialog, QTableWidget, QTableWidgetItem, 
                             QCheckBox, QListWidget, QListWidgetItem)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSettings, QTimer
from PyQt5.QtGui import QIcon
import hashlib

# 注册 sqlite3 的 datetime 适配器和转换器
def adapt_datetime(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def convert_datetime(s):
    s = s.decode('utf-8') if isinstance(s, bytes) else s
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f")

sqlite3.register_adapter(datetime, adapt_datetime)
sqlite3.register_converter("DATETIME", convert_datetime)

# 检查依赖
try:
    from PyQt5.QtWidgets import QApplication
    import pandas as pd
except ImportError as e:
    print(f"缺少必要的依赖库: {e}")
    print("请安装所需库: pip install PyQt5 pandas")
    sys.exit(1)

# 数据库初始化
def init_db():
    conn = sqlite3.connect('shop_system.db', detect_types=sqlite3.PARSE_DECLTYPES)
    cursor = conn.cursor()

    cursor.execute(''' SELECT count(name) FROM sqlite_master WHERE type='table' AND name='users' ''')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'staff',
                last_failed_attempt INTEGER DEFAULT 0,
                attempt_count INTEGER DEFAULT 0,
                balance REAL DEFAULT 0
            )
        ''')
        admin_hash = hashlib.sha256("hyyf123".encode()).hexdigest()
        cursor.execute('''
            INSERT INTO users (username, password_hash, role, balance)
            VALUES (?, ?, ?, ?)
        ''', ('hyyf', admin_hash, 'admin', 0.0))

    cursor.execute(''' SELECT count(name) FROM sqlite_master WHERE type='table' AND name='transactions' ''')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            CREATE TABLE transactions (
                id INTEGER PRIMARY KEY,
                user_id INTEGER,
                type TEXT,
                amount REAL,
                category TEXT,
                timestamp DATETIME,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        ''')
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_trans_user ON transactions(user_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_trans_time ON transactions(timestamp)")

    conn.commit()
    conn.close()

class LoginWindow(QDialog):
    MAX_ATTEMPTS = 3
    LOCK_TIME = 30

    def __init__(self):
        super().__init__()
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setWindowTitle("安全登录")
        self.setFixedSize(550, 450)
        self.attempts = 0
        self.locked = False
        self.settings = QSettings("MyCompany", "ShopSystem")
        self.init_ui()
        self.load_saved_credentials()

    def init_ui(self):
        main_layout = QVBoxLayout()
        title = QLabel('用户消费记账系统')
        title.setStyleSheet('font-size: 36px; font-weight: bold;')
        main_layout.addWidget(title)

        self.username_input = QLineEdit()
        main_layout.addWidget(QLabel("用户名:"))
        main_layout.addWidget(self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        main_layout.addWidget(QLabel("密码:"))
        main_layout.addWidget(self.password_input)

        self.remember_check = QCheckBox("记住我")
        main_layout.addWidget(self.remember_check)

        btn_layout = QHBoxLayout()
        self.btn_login = QPushButton('登录')
        self.btn_login.clicked.connect(self.attempt_login)
        btn_layout.addWidget(self.btn_login)

        self.btn_reset = QPushButton('重置密码')
        self.btn_reset.clicked.connect(self.show_password_reset)
        btn_layout.addWidget(self.btn_reset)

        main_layout.addLayout(btn_layout)
        self.setLayout(main_layout)

    def load_saved_credentials(self):
        username = self.settings.value("username", "")
        password = self.settings.value("password", "")
        remember = self.settings.value("remember", "false") == "true"
        if username and remember:
            self.username_input.setText(username)
            self.password_input.setText(password)
            self.remember_check.setChecked(True)

    def save_credentials(self):
        if self.remember_check.isChecked():
            self.settings.setValue("username", self.username_input.text())
            self.settings.setValue("password", self.password_input.text())
            self.settings.setValue("remember", "true")
        else:
            self.settings.remove("password")
            self.settings.setValue("remember", "false")

    def attempt_login(self):
        if self.locked:
            QMessageBox.warning(self, "账户锁定", f"账户已锁定，请{self.LOCK_TIME}秒后再试")
            return

        username = self.username_input.text().strip()
        password = self.password_input.text()

        if not username or not password:
            QMessageBox.warning(self, "输入错误", "用户名和密码不能为空")
            return

        if self.verify_credentials(username, password):
            self.save_credentials()
            self.accept()
        else:
            self.attempts += 1
            remaining = self.MAX_ATTEMPTS - self.attempts
            if remaining > 0:
                QMessageBox.warning(self, "登录失败", f"用户名或密码错误，剩余尝试次数: {remaining}")
            else:
                self.lock_account()
                QMessageBox.critical(self, "账户锁定", f"超过最大尝试次数，账户已锁定{self.LOCK_TIME}秒")

    def verify_credentials(self, username, password):
        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT password_hash, attempt_count, last_failed_attempt 
                FROM users WHERE username=?
            ''', (username,))
            result = cursor.fetchone()

            if not result:
                return False

            stored_hash, attempts, last_attempt = result
            input_hash = hashlib.sha256(password.encode()).hexdigest()

            if attempts >= self.MAX_ATTEMPTS:
                current_time = int(datetime.now().timestamp())
                if current_time - last_attempt < self.LOCK_TIME:
                    self.locked = True
                    return False
                else:
                    cursor.execute('UPDATE users SET attempt_count=0 WHERE username=?', (username,))
                    conn.commit()

            if stored_hash == input_hash:
                cursor.execute('UPDATE users SET attempt_count=0 WHERE username=?', (username,))
                conn.commit()
                return True
            else:
                cursor.execute('''
                    UPDATE users 
                    SET attempt_count=attempt_count+1, 
                        last_failed_attempt=? 
                    WHERE username=?
                ''', (int(datetime.now().timestamp()), username))
                conn.commit()
                return False

    def lock_account(self):
        self.locked = True
        QTimer.singleShot(self.LOCK_TIME * 1000, self.unlock_account)

    def unlock_account(self):
        self.locked = False
        self.attempts = 0
        QMessageBox.information(self, "账户解锁", "您可以重新尝试登录")

    def show_password_reset(self):
        email, ok = QInputDialog.getText(self, "重置密码", "请输入注册邮箱:")
        if ok and email:
            QMessageBox.information(self, "已发送", "重置链接已发送到您的邮箱（模拟）")

class AsyncLoader(QThread):
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, query, params=()):
        super().__init__()
        self.query = query
        self.params = params

    def run(self):
        try:
            with sqlite3.connect('shop_system.db') as conn:
                cursor = conn.cursor()
                cursor.execute(self.query, self.params)
                self.finished.emit(cursor.fetchall())
        except Exception as e:
            self.error.emit(str(e))

class AddUserDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('新增用户')
        layout = QFormLayout()
        self.name_input = QLineEdit()
        layout.addRow('用户名:', self.name_input)
        btn_confirm = QPushButton('确定')
        btn_confirm.clicked.connect(self.accept)
        layout.addRow(btn_confirm)
        self.setLayout(layout)

class DetailWindow(QWidget):
    balance_updated = pyqtSignal()

    def __init__(self, user_id, parent=None):
        super().__init__(flags=Qt.Window)
        self.setAttribute(Qt.WA_DeleteOnClose)
        self.user_id = user_id
        self.parent = parent

        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT username FROM users WHERE id=?", (self.user_id,))
            username = cursor.fetchone()[0]
            if username == 'hyyf':
                QMessageBox.information(self, "提示", "超级管理员！")
                self.close()
                return

        self.init_ui()
        self.load_data()

    def init_ui(self):
        layout = QVBoxLayout()
        self.lbl_info = QLabel()
        layout.addWidget(self.lbl_info)

        btn_layout = QHBoxLayout()
        self.btn_topup = QPushButton('充值')
        self.btn_consume = QPushButton('消费')
        self.btn_back = QPushButton('返回')
        btn_layout.addWidget(self.btn_topup)
        btn_layout.addWidget(self.btn_consume)
        btn_layout.addWidget(self.btn_back)
        layout.addLayout(btn_layout)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["时间", "类型", "金额", "类目"])
        layout.addWidget(self.table)

        self.lbl_summary = QLabel()
        layout.addWidget(self.lbl_summary)

        self.setLayout(layout)
        self.setMinimumSize(1250, 600)

        self.btn_topup.clicked.connect(self.on_topup_clicked)
        self.btn_consume.clicked.connect(self.on_consume_clicked)
        self.btn_back.clicked.connect(self.close)

    def load_data(self):
        with sqlite3.connect('shop_system.db', detect_types=sqlite3.PARSE_DECLTYPES) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT username, balance FROM users WHERE id=?", (self.user_id,))
            result = cursor.fetchone()
            if result:
                username, balance = result
                self.lbl_info.setText(f"用户信息: {username} (ID: {self.user_id}) 余额: ¥{balance:.2f}")

            cursor.execute("""
                SELECT timestamp, type, amount, category 
                FROM transactions 
                WHERE user_id=? 
                ORDER BY timestamp DESC
            """, (self.user_id,))
            transactions = cursor.fetchall()

            self.table.setRowCount(len(transactions))
            for row, (timestamp, trans_type, amount, category) in enumerate(transactions):
                formatted_timestamp = timestamp.strftime("%Y-%m-%d %H:%M:%S")
                time_item = QTableWidgetItem(formatted_timestamp)
                time_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 0, time_item)

                type_item = QTableWidgetItem(trans_type)
                type_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 1, type_item)

                amount_item = QTableWidgetItem(f"{amount:.2f}")
                amount_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 2, amount_item)

                category_item = QTableWidgetItem(category)
                category_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 3, category_item)

            self.table.resizeColumnsToContents()
            font_metrics = self.table.fontMetrics()
            char_width = font_metrics.averageCharWidth()
            extra_padding = char_width * 12
            for col in range(self.table.columnCount()):
                current_width = self.table.columnWidth(col)
                self.table.setColumnWidth(col, current_width + extra_padding)

            total_income = sum(t[2] for t in transactions if t[1] == 'topup')
            total_expense = sum(t[2] for t in transactions if t[1] == 'consumption')
            self.lbl_summary.setText(f"总收入: ¥{total_income:.2f} 总支出: ¥{total_expense:.2f}")

    def on_topup_clicked(self):
        amount, ok = QInputDialog.getDouble(self, '充值', '输入充值金额:', 0, 0, 1000000, 2)
        if ok and amount > 0:
            self.save_transaction('topup', amount, '充值')

    def on_consume_clicked(self):
        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT balance FROM users WHERE id=?", (self.user_id,))
            balance = cursor.fetchone()[0]

        amount, ok = QInputDialog.getDouble(self, '消费', '输入消费金额:', 0, 0, balance, 2)
        if not ok or amount <= 0:
            return

        if amount > balance:
            QMessageBox.warning(self, "余额不足", f"当前余额: ¥{balance:.2f}，不足以支付 ¥{amount:.2f}")
            return

        category, ok = QInputDialog.getText(self, '消费类目', '请输入消费类目:')
        if not ok or not category.strip():
            QMessageBox.warning(self, "输入错误", "消费类目不能为空")
            return

        self.save_transaction('consumption', amount, category.strip())

    def save_transaction(self, trans_type, amount, category):
        with sqlite3.connect('shop_system.db', detect_types=sqlite3.PARSE_DECLTYPES) as conn:
            cursor = conn.cursor()
            if trans_type == 'topup':
                cursor.execute("UPDATE users SET balance = balance + ? WHERE id=?", (amount, self.user_id))
            else:
                cursor.execute("SELECT balance FROM users WHERE id=?", (self.user_id,))
                balance = cursor.fetchone()[0]
                if amount > balance:
                    QMessageBox.warning(self, "余额不足", f"当前余额: ¥{balance:.2f}，不足以支付 ¥{amount:.2f}")
                    return
                cursor.execute("UPDATE users SET balance = balance - ? WHERE id=?", (amount, self.user_id))
            timestamp = datetime.now().replace(microsecond=0)
            cursor.execute('''
                INSERT INTO transactions (user_id, type, amount, category, timestamp)
                VALUES (?, ?, ?, ?, ?)
            ''', (self.user_id, trans_type, amount, category, timestamp))
            conn.commit()
            self.load_data()
            self.balance_updated.emit()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_page = 1
        self.init_ui()
        self.refresh_users()

    def search_user(self):
        search_text = self.search_input.text().strip()
        if not search_text:
            QMessageBox.warning(self, "输入错误", "请输入用户名或ID")
            return

        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            query = '''
                SELECT id FROM users 
                WHERE id = ? OR username LIKE ?
            '''
            try:
                search_id = int(search_text) if search_text.isdigit() else -1
                cursor.execute(query, (search_id, f'%{search_text}%'))
                result = cursor.fetchone()
            except ValueError:
                cursor.execute(query, (-1, f'%{search_text}%'))
                result = cursor.fetchone()

            if result:
                user_id = result[0]
                self.search_input.clear()
                self.suggestion_list.hide()
                self.show_detail(user_id)
            else:
                QMessageBox.warning(self, "无此用户", f"未找到用户: {search_text}")

    def init_ui(self):
        central_widget = QWidget()
        main_layout = QVBoxLayout()

        top_layout = QHBoxLayout()
        title = QLabel('用户消费记账系统')
        title.setStyleSheet('font-size: 30px; font-weight: bold;')
        top_layout.addWidget(title)

        top_layout.addStretch()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入用户名或ID搜索")
        self.search_input.setFixedWidth(300)
        self.search_input.setStyleSheet("border: 1px solid gray; border-radius: 5px; padding: 2px;")
        self.search_input.textChanged.connect(self.update_search_suggestions)
        top_layout.addWidget(self.search_input)

        self.btn_search = QPushButton('搜索')
        self.btn_search.clicked.connect(self.search_user)
        top_layout.addWidget(self.btn_search)

        main_layout.addLayout(top_layout)

        self.suggestion_list = QListWidget()
        self.suggestion_list.setFixedWidth(600)
        self.suggestion_list.hide()
        self.suggestion_list.itemClicked.connect(self.on_suggestion_clicked)
        self.suggestion_list.setStyleSheet("""
            QListWidget {
                border: 1px solid gray;
                border-radius: 5px;
            }
            QListWidget::item:hover {
                background-color: #e0e0e0;
            }
        """)
        self.suggestion_list.setMaximumHeight(150)
        main_layout.addWidget(self.suggestion_list)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.user_container = QWidget()
        self.user_layout = QVBoxLayout()
        self.user_container.setLayout(self.user_layout)
        self.scroll_area.setWidget(self.user_container)
        main_layout.addWidget(self.scroll_area)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton('创建用户')
        self.btn_revenue = QPushButton('营收情况')
        self.btn_import = QPushButton('导入Excel')
        self.btn_export = QPushButton('导出Excel')
        self.btn_export_summary = QPushButton('导出收支明细')  # 新增按钮
        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_revenue)
        btn_layout.addWidget(self.btn_import)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_export_summary)  # 添加到布局
        main_layout.addLayout(btn_layout)

        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.btn_add.clicked.connect(self.add_user)
        self.btn_revenue.clicked.connect(self.show_revenue)
        self.btn_import.clicked.connect(self.import_from_excel)
        self.btn_export.clicked.connect(self.export_to_excel)
        self.btn_export_summary.clicked.connect(self.export_summary_to_excel)  # 绑定新方法

    def update_search_suggestions(self):
        search_text = self.search_input.text().strip()
        if not search_text:
            self.suggestion_list.hide()
            return

        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            query = '''
                SELECT id, username FROM users 
                WHERE id = ? OR username LIKE ?
            '''
            try:
                search_id = int(search_text) if search_text.isdigit() else -1
                cursor.execute(query, (search_id, f'%{search_text}%'))
                results = cursor.fetchall()
            except ValueError:
                cursor.execute(query, (-1, f'%{search_text}%'))
                results = cursor.fetchall()

        self.suggestion_list.clear()
        if results:
            self.suggestion_list.show()
            for user_id, username in results:
                item = QListWidgetItem(f"{username} (ID: {user_id})")
                item.setData(Qt.UserRole, user_id)
                self.suggestion_list.addItem(item)
        else:
            self.suggestion_list.hide()

    def on_suggestion_clicked(self, item):
        user_id = item.data(Qt.UserRole)
        self.search_input.clear()
        self.suggestion_list.hide()
        self.show_detail(user_id)

    def refresh_users(self, page=1):
        page = max(1, page)
        self.current_page = page
        while self.user_layout.count() > 0:
            item = self.user_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            del item

        loading = QLabel("加载中...")
        self.user_layout.addWidget(loading)

        self.loader = AsyncLoader('''
            SELECT id, username, balance FROM users 
            ORDER BY id LIMIT 10 OFFSET ?
        ''', ((page-1)*10,))
        self.loader.finished.connect(lambda data: self.update_users(page, data))
        self.loader.error.connect(lambda e: QMessageBox.critical(self, '错误', e))
        self.loader.start()

    def update_users(self, page, users):
        while self.user_layout.count() > 0:
            item = self.user_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            del item

        for user_id, username, balance in users:
            item = QWidget()
            layout = QHBoxLayout()
            if username == 'hyyf':
                layout.addWidget(QLabel(f"{username} (ID: {user_id})"))
            else:
                layout.addWidget(QLabel(f"{username} (ID: {user_id})  余额: ¥{balance:.2f}"))
            btn = QPushButton("详情")
            btn.clicked.connect(lambda _, uid=user_id: self.show_detail(uid))
            layout.addWidget(btn)
            item.setLayout(layout)
            self.user_layout.addWidget(item)

        nav = QWidget()
        nav_layout = QHBoxLayout()
        btn_prev = QPushButton("上一页")
        btn_prev.setEnabled(page > 1)
        btn_prev.clicked.connect(lambda: self.refresh_users(page - 1))
        nav_layout.addWidget(btn_prev)

        nav_layout.addStretch()
        page_layout = QHBoxLayout()
        page_input = QLineEdit()
        page_input.setText(str(page))
        page_input.setFixedWidth(50)
        page_input.setAlignment(Qt.AlignCenter)
        page_layout.addWidget(page_input)
        page_layout.addWidget(QLabel(f"/ {self.get_max_pages()} 页"))
        nav_layout.addLayout(page_layout)
        nav_layout.addStretch()

        btn_next = QPushButton("下一页")
        btn_next.clicked.connect(lambda: self.refresh_users(page + 1))
        nav_layout.addWidget(btn_next)

        nav.setLayout(nav_layout)
        self.user_layout.addWidget(nav)

        self.count_loader = AsyncLoader("SELECT COUNT(*) FROM users")
        self.count_loader.finished.connect(
            lambda data: btn_next.setEnabled(page * 10 < data[0][0]))
        self.count_loader.start()

        page_input.returnPressed.connect(lambda: self.jump_to_page(page_input.text()))

    def get_max_pages(self):
        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM users")
            total_users = cursor.fetchone()[0]
        return (total_users + 9) // 10

    def jump_to_page(self, input_text):
        try:
            target_page = int(input_text)
            max_pages = self.get_max_pages()
            if 1 <= target_page <= max_pages:
                self.refresh_users(target_page)
            else:
                QMessageBox.warning(self, "页码错误", "请输入正确的页码！")
        except ValueError:
            QMessageBox.warning(self, "页码错误", "请输入正确的页码！")

    def show_detail(self, user_id):
        self.detail_window = DetailWindow(user_id, self)
        self.detail_window.balance_updated.connect(lambda: self.refresh_users(self.current_page))
        self.detail_window.show()

    def add_user(self):
        dialog = AddUserDialog()
        if dialog.exec_():
            username = dialog.name_input.text().strip()
            if not username:
                QMessageBox.warning(self, "错误", "用户名不能为空")
                return
                
            with sqlite3.connect('shop_system.db') as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM users WHERE username=?", (username,))
                if cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "错误", "该用户名已存在，请使用其他用户名")
                    return
                    
                password_hash = hashlib.sha256("123456".encode()).hexdigest()
                try:
                    cursor.execute('INSERT INTO users (username, password_hash) VALUES (?, ?)', 
                                (username, password_hash))
                    conn.commit()
                    self.refresh_users()
                except sqlite3.Error as e:
                    QMessageBox.critical(self, "错误", f"添加用户失败: {str(e)}")

    def import_from_excel(self):
        from PyQt5.QtWidgets import QFileDialog
        
        # 提示用户确认覆盖操作
        reply = QMessageBox.question(self, "确认覆盖", "导入将覆盖现有数据，是否继续？",
                            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return
        
        # 选择 Excel 文件
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return

        try:
            # 读取 Excel 文件的所有工作表
            excel_file = pd.ExcelFile(file_path)
            imported_users = 0
            imported_transactions = 0
            skipped_hyyf = False  # 专门记录是否跳过了 hyyf
            invalid_usernames = []  # 记录无效用户名（空用户名）

            # 检查 Excel 文件中的重复用户名
            all_usernames = set()
            duplicate_usernames = set()
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                columns = df.columns.tolist()

                if all(col in columns for col in ['username']) and len(columns) == 1:
                    for index, row in df.iterrows():
                        username = str(row['username']).strip()
                        if not username:  # 记录空用户名
                            invalid_usernames.append(f"工作表 '{sheet_name}' 第 {index+2} 行")
                            continue
                        if username in all_usernames:
                            duplicate_usernames.add(username)
                        else:
                            all_usernames.add(username)
                elif all(col in columns for col in ['时间', '类型', '金额', '类目']):
                    username = sheet_name
                    if not username:  # 工作表名为空（理论上不应该发生，但以防万一）
                        invalid_usernames.append(f"工作表 '{sheet_name}'")
                        continue
                    if username in all_usernames:
                        duplicate_usernames.add(username)
                    else:
                        all_usernames.add(username)

            # 如果有重复用户名，提示用户并停止导入
            if duplicate_usernames:
                QMessageBox.critical(self, "导入错误", 
                                    f"Excel 文件中存在重复的用户名：{', '.join(duplicate_usernames)}\n"
                                    "请确保所有用户名唯一后再导入！")
                return

            # 如果有空用户名，提示用户并停止导入
            if invalid_usernames:
                QMessageBox.critical(self, "导入错误", 
                                    f"Excel 文件中存在空用户名：\n" + "\n".join(invalid_usernames) + "\n"
                                    "请确保所有用户名不为空后再导入！")
                return

            # 检查是否包含管理员用户 hyyf
            if 'hyyf' in all_usernames:
                skipped_hyyf = True

            with sqlite3.connect('shop_system.db', detect_types=sqlite3.PARSE_DECLTYPES) as conn:
                cursor = conn.cursor()

                # 清空数据库
                cursor.execute("DELETE FROM transactions")
                cursor.execute("DELETE FROM users")

                # 重置 users 表的自增计数器
                cursor.execute("DELETE FROM sqlite_sequence WHERE name='users'")

                # 重新插入管理员用户
                admin_hash = hashlib.sha256("hyyf123".encode()).hexdigest()
                cursor.execute('''
                    INSERT INTO users (username, password_hash, role, balance)
                    VALUES (?, ?, ?, ?)
                ''', ('hyyf', admin_hash, 'admin', 0.0))

                # 定义支持的格式
                user_only_format = ['username']
                full_transaction_format = ['时间', '类型', '金额', '类目']

                # 遍历所有工作表
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    columns = df.columns.tolist()

                    # 检查格式 1：仅用户信息
                    if all(col in columns for col in user_only_format) and len(columns) == 1:
                        for index, row in df.iterrows():
                            username = str(row['username']).strip()
                            if username == 'hyyf':  # 跳过管理员用户
                                continue
                            
                            password_hash = hashlib.sha256("123456".encode()).hexdigest()
                            cursor.execute('INSERT INTO users (username, password_hash) VALUES (?, ?)',
                                        (username, password_hash))
                            imported_users += 1

                    # 检查格式 2：完整交易记录
                    elif all(col in columns for col in full_transaction_format):
                        username = sheet_name
                        if username == 'hyyf':  # 跳过管理员用户
                            continue

                        password_hash = hashlib.sha256("123456".encode()).hexdigest()
                        cursor.execute('INSERT INTO users (username, password_hash) VALUES (?, ?)',
                                    (username, password_hash))
                        user_id = cursor.lastrowid
                        imported_users += 1

                        for index, row in df.iterrows():
                            timestamp = row['时间']
                            trans_type = row['类型']
                            amount = float(row['金额'])
                            category = row['类目']

                            # 处理 timestamp，确保它是 datetime 对象
                            if pd.isna(timestamp):  # 检查空值
                                QMessageBox.critical(self, "时间为空", f"工作表 '{sheet_name}' 第 {index+2} 行的 '时间' 列为空")
                                return
                            elif isinstance(timestamp, pd.Timestamp):
                                # 直接将 pandas.Timestamp 转换为 datetime
                                timestamp = timestamp.to_pydatetime()
                            elif isinstance(timestamp, str):
                                try:
                                    # 尝试将字符串解析为 datetime
                                    timestamp = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                                except ValueError:
                                    try:
                                        # 如果格式不匹配，尝试带毫秒的格式
                                        timestamp = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")
                                    except ValueError as e:
                                        QMessageBox.critical(self, "时间格式错误", f"工作表 '{sheet_name}' 第 {index+2} 行的时间格式不正确: {timestamp}\n请使用 'YYYY-MM-DD HH:MM:SS' 格式")
                                        return
                            elif isinstance(timestamp, (int, float)):
                                # 如果是 Excel 日期序列号，转换为 datetime
                                try:
                                    timestamp = pd.Timestamp("1899-12-30") + pd.Timedelta(days=timestamp)
                                    timestamp = timestamp.to_pydatetime()
                                except Exception as e:
                                    QMessageBox.critical(self, "时间格式错误", f"工作表 '{sheet_name}' 第 {index+2} 行的日期序列号转换失败: {timestamp}")
                                    return
                            elif isinstance(timestamp, datetime):
                                # 如果已经是 datetime 对象，直接使用
                                pass
                            else:
                                # 如果是其他类型，抛出错误
                                QMessageBox.critical(self, "时间格式错误", f"工作表 '{sheet_name}' 第 {index+2} 行的日期格式不受支持: {timestamp} (类型: {type(timestamp)})")
                                return

                            # 插入交易记录
                            cursor.execute('''
                                INSERT INTO transactions (user_id, type, amount, category, timestamp)
                                VALUES (?, ?, ?, ?, ?)
                            ''', (user_id, trans_type, amount, category, timestamp))
                            
                            if trans_type == 'topup':
                                cursor.execute("UPDATE users SET balance = balance + ? WHERE id=?", (amount, user_id))
                            elif trans_type == 'consumption':
                                cursor.execute("UPDATE users SET balance = balance - ? WHERE id=?", (amount, user_id))
                            
                            imported_transactions += 1

                    else:
                        QMessageBox.critical(self, "格式错误", 
                                        f"工作表 '{sheet_name}' 格式不正确。\n"
                                        "支持以下两种格式：\n"
                                        "1. 仅包含 'username' 列（导入用户）\n"
                                        "2. 包含 '时间', '类型', '金额', '类目' 列（与导出格式一致，导入交易记录）")
                        return

                conn.commit()

            self.refresh_users()
            # 改进提示信息
            if skipped_hyyf:
                QMessageBox.information(self, "导入完成", 
                                    f"成功导入 {imported_users} 个用户，{imported_transactions} 条交易记录，\n"
                                    f"跳过了管理员用户 'hyyf'")
            else:
                QMessageBox.information(self, "导入完成", 
                                    f"成功导入 {imported_users} 个用户，{imported_transactions} 条交易记录")

        except ValueError as ve:
            QMessageBox.critical(self, "数据错误", f"数据格式错误，请检查金额是否为数字: {str(ve)}")
        except sqlite3.IntegrityError as ie:
            QMessageBox.critical(self, "导入失败", f"数据库错误：{str(ie)}\n可能是用户名重复导致！")
        except Exception as e:
            QMessageBox.critical(self, "导入错误", f"导入失败: {str(e)}")

    def export_to_excel(self):
        def on_export_finished():
            QMessageBox.information(self, "完成", "交易记录导出成功")
            self.btn_export.setEnabled(True)

        self.btn_export.setEnabled(False)
        self.export_thread = AsyncLoader("SELECT id, username FROM users")
        self.export_thread.finished.connect(
            lambda users: self._export_users(users, on_export_finished))
        self.export_thread.start()

    def _export_users(self, users, callback):
        try:
            from openpyxl.utils import get_column_letter
            print(f"Exporting {len(users)} users")

            with sqlite3.connect('shop_system.db', detect_types=sqlite3.PARSE_DECLTYPES) as conn:
                with pd.ExcelWriter('交易记录.xlsx', engine='openpyxl') as writer:
                    for user_id, username in users:
                        print(f"Exporting transactions for user: {username}")
                        for chunk in pd.read_sql(
                            f'''SELECT 
                                timestamp as 时间,
                                type as 类型,
                                amount as 金额,
                                category as 类目
                            FROM transactions 
                            WHERE user_id={user_id}
                            ORDER BY timestamp DESC''',
                            conn, chunksize=1000
                        ):
                            chunk.to_excel(writer, sheet_name=username[:30], index=False)

                    # 调整列宽
                    workbook = writer.book
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for column in sheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            for cell in column:
                                if cell.value:
                                    cell_value = str(cell.value)
                                    length = sum(2 if ord(char) > 127 else 1 for char in cell_value)
                                    max_length = max(max_length, length)
                            adjusted_width = max_length + 2
                            sheet.column_dimensions[column_letter].width = adjusted_width

            print("Transaction export completed")
            callback()

        except PermissionError:
            print("PermissionError: 无法写入文件")
            QMessageBox.critical(self, "导出错误", "无法导出，请关闭‘交易记录.xlsx’文件后再试！")
            self.btn_export.setEnabled(True)
        except Exception as e:
            print(f"Export error: {str(e)}")
            QMessageBox.critical(self, "导出错误", f"导出失败: {str(e)}")
            self.btn_export.setEnabled(True)

    def export_summary_to_excel(self):
        def on_export_summary_finished():
            QMessageBox.information(self, "完成", "收支明细导出成功")
            self.btn_export_summary.setEnabled(True)

        self.btn_export_summary.setEnabled(False)
        self.export_summary_thread = AsyncLoader("SELECT id, username FROM users")
        self.export_summary_thread.finished.connect(
            lambda users: self._export_summary(users, on_export_summary_finished))
        self.export_summary_thread.start()

    def _export_summary(self, users, callback):
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Alignment, Border, Side  # 引入 Border 和 Side 用于设置边框
            print(f"Exporting summary for {len(users)} users")

            # 准备收支明细数据
            with sqlite3.connect('shop_system.db', detect_types=sqlite3.PARSE_DECLTYPES) as conn:
                cursor = conn.cursor()
                data = []
                for user_id, username in users:
                    cursor.execute("SELECT SUM(amount) FROM transactions WHERE user_id=? AND type='topup'", (user_id,))
                    total_topup = cursor.fetchone()[0] or 0.0
                    cursor.execute("SELECT SUM(amount) FROM transactions WHERE user_id=? AND type='consumption'", (user_id,))
                    total_consumption = cursor.fetchone()[0] or 0.0
                    cursor.execute("SELECT balance FROM users WHERE id=?", (user_id,))
                    balance = cursor.fetchone()[0] or 0.0
                    data.append([username, total_topup, total_consumption, balance])

            # 调试：打印 data 内容，确认是否有数据
            print(f"Data for summary: {data}")

            # 创建收支明细 DataFrame，添加“序号”列
            if not data:
                data = [['无数据', 0.0, 0.0, 0.0]]
            # 为每个用户添加序号
            numbered_data = [[i + 1] + row for i, row in enumerate(data)]
            summary_df = pd.DataFrame(numbered_data, columns=['序号', '姓名', '充值', '消费', '余额'])

            # 计算总计
            total_topup_sum = sum(row[1] for row in data if row[0] != '无数据')
            total_consumption_sum = sum(row[2] for row in data if row[0] != '无数据')
            total_balance_sum = sum(row[3] for row in data if row[0] != '无数据')
            # 调试：打印总计值，确认计算结果
            print(f"Total topup: {total_topup_sum}, Total consumption: {total_consumption_sum}, Total balance: {total_balance_sum}")

            # 如果总计值全为 0，说明没有交易记录，添加提示
            if total_topup_sum == 0 and total_consumption_sum == 0 and total_balance_sum == 0:
                print("No transactions found for summary.")

            # 总计行：将“总充值”等放在 B 列，计算值放在 C 列，D、E 列为空
            total_row1 = ['', '总充值', total_topup_sum, '', '']
            total_row2 = ['', '总消费', total_consumption_sum, '', '']
            total_row3 = ['', '总余额', total_balance_sum, '', '']
            total_df = pd.DataFrame([total_row1, total_row2, total_row3], columns=['序号', '姓名', '充值', '消费', '余额'])

            # 写入收支明细到单独的 Excel 文件
            with pd.ExcelWriter('收支明细.xlsx', engine='openpyxl') as writer:
                # 写入用户数据，空两行后写入总计
                summary_df.to_excel(writer, sheet_name='收支明细', index=False, startrow=0)
                # 设置 header=False，避免重复写入列标题
                total_df.to_excel(writer, sheet_name='收支明细', index=False, header=False, startrow=len(summary_df) + 2)

                # 获取工作簿和工作表
                workbook = writer.book
                sheet = workbook['收支明细']

                # 设置“序号”列（A 列）居中
                for row in range(1, len(summary_df) + 6):  # 覆盖标题行、用户数据行和总计行
                    cell = sheet.cell(row=row, column=1)  # A 列
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 设置所有计算值（C、D、E 列）居中
                for row in range(1, len(summary_df) + 2):  # 覆盖标题行和用户数据行
                    for col in range(3, 6):  # C、D、E 列（3到5）
                        cell = sheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 设置总计行的计算值（C、D、E 列合并后居中）
                for row in range(len(summary_df) + 3, len(summary_df) + 6):  # 总计行的范围
                    # 合并 C、D、E 列
                    sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)  # 合并 C、D、E 列
                    # 设置合并后的单元格居中
                    cell = sheet.cell(row=row, column=3)  # C 列（合并后的起始列）
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 每隔一行填充浅灰色背景（从第2行开始，第1行是标题行）
                fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # 浅灰色
                for row in range(2, len(summary_df) + 2):  # 从第2行（数据行）开始
                    if (row - 2) % 2 == 1:  # 每隔一行（奇数行）
                        for col in range(1, 6):  # A到E列（1到5）
                            cell = sheet.cell(row=row, column=col)
                            cell.fill = fill

                # 为所有单元格添加边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in range(1, len(summary_df) + 6):  # 覆盖标题行、用户数据行和总计行
                    for col in range(1, 6):  # A到E列（1到5）
                        cell = sheet.cell(row=row, column=col)
                        cell.border = thin_border

                # 调整列宽，直接设置为目标值，不受内容长度影响
                column_widths = {'A': 5.5, 'B': 40, 'C': 13, 'D': 13, 'E': 13}  # 总宽度 5.5+40+13+13+13=84.5
                for column in sheet.columns:
                    column_letter = get_column_letter(column[0].column)
                    # 直接设置目标宽度
                    sheet.column_dimensions[column_letter].width = column_widths[column_letter]

            print("Summary export completed")
            callback()

        except PermissionError:
            print("PermissionError: 无法写入文件")
            QMessageBox.critical(self, "导出错误", "无法导出，请关闭‘收支明细.xlsx’文件后再试！")
            self.btn_export_summary.setEnabled(True)
        except Exception as e:
            print(f"Summary export error: {str(e)}")
            QMessageBox.critical(self, "导出错误", f"导出失败: {str(e)}")
            self.btn_export_summary.setEnabled(True)

    def calculate_revenue(self):
        with sqlite3.connect('shop_system.db') as conn:
            cursor = conn.cursor()
            
            # 计算所有用户的充值总额
            cursor.execute("""
                SELECT SUM(amount) 
                FROM transactions 
                WHERE type = 'topup'
            """)
            total_topup = cursor.fetchone()[0] or 0.0
            
            # 计算所有用户的消费总额
            cursor.execute("""
                SELECT SUM(amount) 
                FROM transactions 
                WHERE type = 'consumption'
            """)
            total_consumption = cursor.fetchone()[0] or 0.0
            
            # 计算所有用户的余额总和
            cursor.execute("SELECT SUM(balance) FROM users WHERE id != 1")
            total_balance = cursor.fetchone()[0] or 0.0

            # 统计用户数量
            cursor.execute("SELECT COUNT(*) FROM users WHERE id != 1")
            user_count = cursor.fetchone()[0] or 0

        return total_topup, total_consumption, total_balance, user_count
    
    def show_revenue(self):
        total_topup, total_consumption, total_balance, user_count = self.calculate_revenue()
        
        revenue_dialog = QMessageBox(self)
        revenue_dialog.setWindowTitle("营收情况")
        revenue_dialog.setText(
            f"用户数量：{user_count} 位\n"
            f"用户充值：{total_topup:.2f} 元\n"
            f"用户消费：{total_consumption:.2f} 元\n"
            f"余额：{total_balance:.2f} 元"
        )
        revenue_dialog.setStandardButtons(QMessageBox.Ok)
        revenue_dialog.exec_()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    init_db()
    login = LoginWindow()
    if login.exec_() == QDialog.Accepted:
        window = MainWindow()
        window.resize(1250, 1300)
        window.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)